# -*- coding: utf-8 -*-
"""
Created on Wed Feb 15 14:26:52 2023

@author: Admin
"""

from datetime import datetime
import openpyxl
import re
import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt





enter=input('輸入型號或預算:')

    

if str(enter) in '今日價格行情狀況比價':
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64 ) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36'}
    url='https://www.jyes.com.tw/product.php'
    ans=list()
    webpage=requests.get(url,headers=HEADERS)
    soup=BeautifulSoup(webpage.text,'html.parser')
    div=soup.find('div',{'class':'list'})
    tr=div.find_all('div',{'class':'brand-wrap'})
    #時間
    result = datetime.now().strftime("%Y-%m-%d %H:%M:%S ")
    print("更新時間:",result)
    counter=0

    #過濾不要的資料
    for i in tr:   
        if 'TOUGHER' in str(i.find('span','brand')):
            continue
        if 'Décent' in str(i.find('span','brand')):
            continue
        if 'HUAWEI' in str(i.find('span','brand')):
            continue
        if '任天堂Switch' in str(i.find('span','brand')):
            continue

        #從網站找出需要的訊息文字
        alist=i.find_all('td',{'class':'sn'})
        price1=i.find_all('td',{'data-title':'原廠建議售價 :'})
        price2=i.find_all('td',{'data-title':'門市破盤價 :'})
        
        #過濾不要的資料
        for a,b,c in zip(alist,price1,price2):
            if '耳機' in str(a.find('a','tag-link')):
                continue
            if '手錶' in str(a.find('a','tag-link')):
                continue
            if '手環' in str(a.find('a','tag-link')):
                continue
            if '平板' in str(a.find('a','tag-link')):
                continue
            if 'Watch' in str(a.find('a','tag-link')):
                continue
            if '音箱' in str(a.find('a','tag-link')):
                continue
            if 'AirTag' in str(a.find('a','tag-link')):
                continue
            if 'Pencil' in str(a.find('a','tag-link')):
                continue
            if 'TV' in str(a.find('a','tag-link')):
                continue
            if '送' in str(a.find('i','tag-summary')):
                continue
            if '抽' in str(a.find('i','tag-summary')):
                continue
            if 'Chromecast' in str(a.find('a','tag-link')):
                continue
            
            #過濾文字
            d=re.findall('\d+', str(b))#正規表達式選取"數字"
            e=re.findall('\d+', str(c))#正規表達式選取"數字"
            f=int("".join(d))#原價
            g=int("".join(e))#公允價值
            h=('{:.2%}'.format((g-f)/f))#帳跌
            end=(a.text.strip(),'原價:',b.text.strip(),'市價:',c.text.strip(),'漲跌幅度:',h)
            ans.append(end)
            
            #給予條件篩選
            print(a.text.strip(),'原價:',b.text.strip(),'市價:',c.text.strip(),'漲跌幅度:',h)
                           
    #把數據資料寫入EXCEL            
    try:
        wb = openpyxl.load_workbook('手機價格.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        
    # 新建一個工作表，使用當前日期作為工作表名稱
    new_sheet_name = datetime.today().strftime('%Y-%m-%d')
    ws = wb.create_sheet(title=new_sheet_name)
    
    # 標題行
    headers = ['機型','','原價','','市價','','漲跌']
    for j, header in enumerate(headers):
        ws.cell(row=1, column=j+1, value=header)
        
    # 將結果寫入工作表
    for i, row in enumerate(ans):
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)
    
    # 將Excel文件保存
    wb.save('手機價格.xlsx')
    
    

else:
    
    #爬蟲
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64 ) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36'}
    url='https://www.jyes.com.tw/product.php'
    ans=list()
    webpage=requests.get(url,headers=HEADERS)
    soup=BeautifulSoup(webpage.text,'html.parser')
    div=soup.find('div',{'class':'list'})
    tr=div.find_all('div',{'class':'brand-wrap'})
    #時間
    result = datetime.now().strftime("%Y-%m-%d %H:%M:%S ")
    print("更新時間:",result)
    counter=0
    
    #過濾不要的資料
    for i in tr:   
        if 'TOUGHER' in str(i.find('span','brand')):
            continue
        if 'Décent' in str(i.find('span','brand')):
            continue
        if 'HUAWEI' in str(i.find('span','brand')):
            continue
        if '任天堂Switch' in str(i.find('span','brand')):
            continue
    
        #從網站找出需要的訊息文字
        alist=i.find_all('td',{'class':'sn'})
        price1=i.find_all('td',{'data-title':'原廠建議售價 :'})
        price2=i.find_all('td',{'data-title':'門市破盤價 :'})
        
        #過濾不要的資料
        for a,b,c in zip(alist,price1,price2):
            if '耳機' in str(a.find('a','tag-link')):
                continue
            if '手錶' in str(a.find('a','tag-link')):
                continue
            if '手環' in str(a.find('a','tag-link')):
                continue
            if '平板' in str(a.find('a','tag-link')):
                continue
            if 'Watch' in str(a.find('a','tag-link')):
                continue
            if '音箱' in str(a.find('a','tag-link')):
                continue
            if 'AirTag' in str(a.find('a','tag-link')):
                continue
            if 'Pencil' in str(a.find('a','tag-link')):
                continue
            if 'TV' in str(a.find('a','tag-link')):
                continue
            if '送' in str(a.find('i','tag-summary')):
                continue
            if '抽' in str(a.find('i','tag-summary')):
                continue
            if 'Chromecast' in str(a.find('a','tag-link')):
                continue
            
            #過濾文字
            d=re.findall('\d+', str(b))#正規表達式選取"數字"
            e=re.findall('\d+', str(c))#正規表達式選取"數字"
            f=int("".join(d))#原價
            g=int("".join(e))#公允價值
            h=('{:.2%}'.format((g-f)/f))#帳跌
            end=(a.text.strip(),'原價:',b.text.strip(),'市價:',c.text.strip(),'漲跌幅度:',h)
            ans.append(end)
            
            #給予條件篩選
            if enter.isdigit():#預算
                if g<=int(enter):
                    print(a.text.strip(),'原價:',b.text.strip(),'市價:',c.text.strip(),'漲跌幅度:',h)
                    counter+=1
                
            else:#型號
               if enter in str(a) :
                   print(a.text.strip(),'原價:',b.text.strip(),'市價:',c.text.strip(),'漲跌幅度:',h)
                   counter+=1
               
              
    if counter==0 and enter.isdigit():#預算不足
        print('您的預算不足')
    elif counter==0:#找不到型號
        print('此款手機並無販售')
                   
    # 把數據資料寫入EXCEL            
    # 開啟現有的Excel文件，如果不存在就新建一個
    try:
        wb = openpyxl.load_workbook('手機價格.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        
    # 新建一個工作表，使用當前日期作為工作表名稱
    new_sheet_name = datetime.today().strftime('%Y-%m-%d')
    ws = wb.create_sheet(title=new_sheet_name)
    
    # 標題行
    headers = ['機型','','原價','','市價','','漲跌']
    for j, header in enumerate(headers):
        ws.cell(row=1, column=j+1, value=header)
        
    # 將結果寫入工作表
    for i, row in enumerate(ans):
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)
    
    # 將Excel文件保存
    wb.save('手機價格.xlsx')
    
    
    
    
    
    